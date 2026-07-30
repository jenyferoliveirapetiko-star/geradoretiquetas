import base64
import io
import os
import sqlite3
from datetime import datetime
from functools import wraps
from pathlib import Path

import barcode
from barcode.writer import SVGWriter
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
DATABASE_PATH = BASE_DIR / "database" / "database.db"
UPLOAD_FOLDER = BASE_DIR / "uploads"

DATABASE_PATH.parent.mkdir(exist_ok=True)
UPLOAD_FOLDER.mkdir(exist_ok=True)


app = Flask(__name__)

app.config["SECRET_KEY"] = os.getenv(
    "SECRET_KEY",
    "troque-esta-chave-antes-de-publicar",
)

app.config["UPLOAD_FOLDER"] = str(UPLOAD_FOLDER)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024


USERS = {
    "petiko": {
        "empresa": "Petiko",
        "senha": generate_password_hash(
            os.getenv("PETIKO_PASSWORD", "azul2026")
        ),
    },
    "paws": {
        "empresa": "Paws",
        "senha": generate_password_hash(
            os.getenv("PAWS_PASSWORD", "amarelo2026")
        ),
    },
    "innova": {
        "empresa": "Innova",
        "senha": generate_password_hash(
            os.getenv("INNOVA_PASSWORD", "rosa2026")
        ),
    },
}


def conectar_banco():
    conexao = sqlite3.connect(DATABASE_PATH)
    conexao.row_factory = sqlite3.Row
    return conexao


def criar_banco():
    with conectar_banco() as conexao:
        conexao.execute(
            """
            CREATE TABLE IF NOT EXISTS produtos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                empresa TEXT NOT NULL,
                nome TEXT NOT NULL,
                ean TEXT NOT NULL,
                criado_em TEXT NOT NULL,
                UNIQUE (empresa, ean)
            )
            """
        )

        conexao.execute(
            """
            CREATE TABLE IF NOT EXISTS historico (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                empresa TEXT NOT NULL,
                produto TEXT NOT NULL,
                ean TEXT NOT NULL,
                quantidade INTEGER NOT NULL,
                impresso_em TEXT NOT NULL
            )
            """
        )

        conexao.commit()


def login_obrigatorio(funcao):
    @wraps(funcao)
    def protegida(*args, **kwargs):
        if "usuario" not in session:
            flash("Entre com seu usuário e senha.", "erro")
            return redirect(url_for("login"))

        return funcao(*args, **kwargs)

    return protegida


def normalizar_cabecalho(valor):
    if valor is None:
        return ""

    texto = str(valor).strip().lower()

    substituicoes = {
        " ": "",
        "/": "",
        "\\": "",
        "-": "",
        "_": "",
        ".": "",
    }

    for antigo, novo in substituicoes.items():
        texto = texto.replace(antigo, novo)

    return texto


def normalizar_ean(valor):
    if valor is None:
        return ""

    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    return "".join(caractere for caractere in texto if caractere.isdigit())


def arquivo_permitido(nome):
    return "." in nome and nome.rsplit(".", 1)[1].lower() in {"xlsx", "xlsm"}


def gerar_codigo_barras(ean):
    memoria = io.BytesIO()

    opcoes = {
    "module_width": 0.30,
    "module_height": 9,
    "quiet_zone": 1,
    "font_size": 0,
    "text_distance": 0,
    "write_text": False,
}

    try:
        if len(ean) == 13:
            classe = barcode.get_barcode_class("ean13")
            codigo = classe(ean[:12], writer=SVGWriter())

        elif len(ean) == 8:
            classe = barcode.get_barcode_class("ean8")
            codigo = classe(ean[:7], writer=SVGWriter())

        else:
            classe = barcode.get_barcode_class("code128")
            codigo = classe(ean, writer=SVGWriter())

        codigo.write(memoria, options=opcoes)

    except Exception:
        classe = barcode.get_barcode_class("code128")
        codigo = classe(ean, writer=SVGWriter())
        codigo.write(memoria, options=opcoes)

    conteudo = base64.b64encode(memoria.getvalue()).decode("utf-8")

    return f"data:image/svg+xml;base64,{conteudo}"


@app.route("/login", methods=["GET", "POST"])
def login():
    if "usuario" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip().lower()
        senha = request.form.get("senha", "")

        cadastro = USERS.get(usuario)

        if cadastro and check_password_hash(cadastro["senha"], senha):
            session.clear()
            session["usuario"] = usuario
            session["empresa"] = cadastro["empresa"]

            flash(
                f"Bem-vindo ao sistema da {cadastro['empresa']}!",
                "sucesso",
            )

            return redirect(url_for("dashboard"))

        flash("Usuário ou senha incorretos.", "erro")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Você saiu do sistema.", "sucesso")
    return redirect(url_for("login"))


@app.route("/")
@login_obrigatorio
def dashboard():
    empresa = session["empresa"]
    pesquisa = request.args.get("pesquisa", "").strip()

    with conectar_banco() as conexao:
        if pesquisa:
            produtos = conexao.execute(
                """
                SELECT id, nome, ean, criado_em
                FROM produtos
                WHERE empresa = ?
                AND (
                    nome LIKE ?
                    OR ean LIKE ?
                )
                ORDER BY nome
                """,
                (
                    empresa,
                    f"%{pesquisa}%",
                    f"%{pesquisa}%",
                ),
            ).fetchall()

        else:
            produtos = conexao.execute(
                """
                SELECT id, nome, ean, criado_em
                FROM produtos
                WHERE empresa = ?
                ORDER BY nome
                """,
                (empresa,),
            ).fetchall()

        quantidade_produtos = conexao.execute(
            """
            SELECT COUNT(*) AS total
            FROM produtos
            WHERE empresa = ?
            """,
            (empresa,),
        ).fetchone()["total"]

        etiquetas_hoje = conexao.execute(
            """
            SELECT COALESCE(SUM(quantidade), 0) AS total
            FROM historico
            WHERE empresa = ?
            AND DATE(impresso_em) = DATE('now', 'localtime')
            """,
            (empresa,),
        ).fetchone()["total"]

        historico = conexao.execute(
            """
            SELECT produto, ean, quantidade, impresso_em
            FROM historico
            WHERE empresa = ?
            ORDER BY id DESC
            LIMIT 10
            """,
            (empresa,),
        ).fetchall()

    return render_template(
        "dashboard.html",
        empresa=empresa,
        produtos=produtos,
        pesquisa=pesquisa,
        quantidade_produtos=quantidade_produtos,
        etiquetas_hoje=etiquetas_hoje,
        historico=historico,
    )


@app.route("/importar", methods=["POST"])
@login_obrigatorio
def importar():
    arquivo = request.files.get("planilha")

    if not arquivo or not arquivo.filename:
        flash("Selecione uma planilha.", "erro")
        return redirect(url_for("dashboard"))

    if not arquivo_permitido(arquivo.filename):
        flash("Envie uma planilha no formato XLSX ou XLSM.", "erro")
        return redirect(url_for("dashboard"))

    nome_seguro = secure_filename(arquivo.filename)
    caminho = UPLOAD_FOLDER / nome_seguro
    arquivo.save(caminho)

    try:
        workbook = load_workbook(caminho, data_only=True)
        planilha = workbook.active

        primeira_linha = next(planilha.iter_rows(min_row=1, max_row=1))

        cabecalhos = {
            normalizar_cabecalho(celula.value): indice
            for indice, celula in enumerate(primeira_linha)
        }

        nomes_produto = {
            "item",
            "iten",
            "produto",
            "nome",
            "nomeproduto",
        }

        nomes_ean = {
            "ean",
            "gtin",
            "gtinean",
            "eangtin",
            "codigodebarras",
            "codigo",
        }

        coluna_produto = next(
            (
                indice
                for nome, indice in cabecalhos.items()
                if nome in nomes_produto
            ),
            None,
        )

        coluna_ean = next(
            (
                indice
                for nome, indice in cabecalhos.items()
                if nome in nomes_ean
            ),
            None,
        )

        if coluna_produto is None or coluna_ean is None:
            flash(
                "Não encontrei as colunas de produto e EAN. "
                "Use os cabeçalhos ITEM e GTIN/EAN.",
                "erro",
            )

            return redirect(url_for("dashboard"))

        empresa = session["empresa"]
        adicionados = 0
        atualizados = 0
        ignorados = 0

        with conectar_banco() as conexao:
            for linha in planilha.iter_rows(min_row=2, values_only=True):
                nome = linha[coluna_produto]
                ean = normalizar_ean(linha[coluna_ean])

                if nome is None or not str(nome).strip() or not ean:
                    ignorados += 1
                    continue

                nome = str(nome).strip()

                existente = conexao.execute(
                    """
                    SELECT id
                    FROM produtos
                    WHERE empresa = ? AND ean = ?
                    """,
                    (empresa, ean),
                ).fetchone()

                if existente:
                    conexao.execute(
                        """
                        UPDATE produtos
                        SET nome = ?
                        WHERE id = ?
                        """,
                        (nome, existente["id"]),
                    )

                    atualizados += 1

                else:
                    conexao.execute(
                        """
                        INSERT INTO produtos (
                            empresa,
                            nome,
                            ean,
                            criado_em
                        )
                        VALUES (?, ?, ?, ?)
                        """,
                        (
                            empresa,
                            nome,
                            ean,
                            datetime.now().isoformat(
                                sep=" ",
                                timespec="seconds",
                            ),
                        ),
                    )

                    adicionados += 1

            conexao.commit()

        flash(
            f"Importação concluída: {adicionados} adicionados, "
            f"{atualizados} atualizados e {ignorados} ignorados.",
            "sucesso",
        )

    except Exception as erro:
        flash(f"Não foi possível importar a planilha: {erro}", "erro")

    finally:
        try:
            caminho.unlink(missing_ok=True)
        except OSError:
            pass

    return redirect(url_for("dashboard"))


@app.route("/produto/<int:produto_id>/excluir", methods=["POST"])
@login_obrigatorio
def excluir_produto(produto_id):
    empresa = session["empresa"]

    with conectar_banco() as conexao:
        produto = conexao.execute(
            """
            SELECT nome
            FROM produtos
            WHERE id = ? AND empresa = ?
            """,
            (produto_id, empresa),
        ).fetchone()

        if not produto:
            flash("Produto não encontrado.", "erro")
            return redirect(url_for("dashboard"))

        conexao.execute(
            """
            DELETE FROM produtos
            WHERE id = ? AND empresa = ?
            """,
            (produto_id, empresa),
        )

        conexao.commit()

    flash(f'Produto "{produto["nome"]}" excluído.', "sucesso")
    return redirect(url_for("dashboard"))
@app.route("/gerar-etiquetas", methods=["POST"])
@login_obrigatorio
def gerar_etiquetas():
    empresa = session["empresa"]

    # Recebe todos os produtos marcados na tela.
    produtos_ids = request.form.getlist("produtos_selecionados")

    if not produtos_ids:
        flash("Selecione pelo menos um produto.", "erro")
        return redirect(url_for("dashboard"))

    etiquetas = []
    produtos_processados = 0
    total_etiquetas = 0

    with conectar_banco() as conexao:
        for produto_id_texto in produtos_ids:
            try:
                produto_id = int(produto_id_texto)
            except (TypeError, ValueError):
                continue

            # Cada produto possui seu próprio campo de quantidade.
            quantidade = request.form.get(
                f"quantidade_{produto_id}",
                type=int,
            )

            if not quantidade or quantidade < 1 or quantidade > 5000:
                flash(
                    "Todas as quantidades devem estar entre 1 e 5.000.",
                    "erro",
                )
                return redirect(url_for("dashboard"))

            produto = conexao.execute(
                """
                SELECT id, nome, ean
                FROM produtos
                WHERE id = ? AND empresa = ?
                """,
                (produto_id, empresa),
            ).fetchone()

            if not produto:
                continue

            codigo_barras = gerar_codigo_barras(produto["ean"])

            # Adiciona cada etiqueta individualmente à lista de impressão.
            for _ in range(quantidade):
                etiquetas.append(
                    {
                        "nome": produto["nome"],
                        "ean": produto["ean"],
                        "codigo_barras": codigo_barras,
                    }
                )

            conexao.execute(
                """
                INSERT INTO historico (
                    empresa,
                    produto,
                    ean,
                    quantidade,
                    impresso_em
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    empresa,
                    produto["nome"],
                    produto["ean"],
                    quantidade,
                    datetime.now().isoformat(
                        sep=" ",
                        timespec="seconds",
                    ),
                ),
            )

            produtos_processados += 1
            total_etiquetas += quantidade

        conexao.commit()

    if not etiquetas:
        flash("Nenhum produto válido foi selecionado.", "erro")
        return redirect(url_for("dashboard"))

    return render_template(
        "etiquetas.html",
        empresa=empresa,
        etiquetas=etiquetas,
        produtos_processados=produtos_processados,
        total_etiquetas=total_etiquetas,
    )
@app.errorhandler(413)
def arquivo_grande(_erro):
    flash("O arquivo é muito grande. O limite é 10 MB.", "erro")
    return redirect(url_for("dashboard"))


criar_banco()


if __name__ == "__main__":
    print("\nGerador de Etiquetas")
    print("Abra no navegador: http://127.0.0.1:5000\n")

    app.run(
        host="127.0.0.1",
        port=5000,
        debug=True,
        )
 